"""Excel helpers — the shop's palette + sheet conventions, semantics ported
from timMtesting/t11_excel_export.py sheet()/paint() and K12Intel
k12/excel_export.py (HOT green / warm amber / act-now red / navy header;
rep-facing tabs first, Methodology & Sources last; money format '"$"#,##0')."""
from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

FILL_HDR = PatternFill("solid", start_color="1F4E79")
FONT_HDR = Font(color="FFFFFF", bold=True)
FILLS = {
    "hot": (PatternFill("solid", start_color="C6EFCE"), Font(color="006100", bold=True)),
    "warm": (PatternFill("solid", start_color="FFEB9C"), Font(color="9C6500")),
    "now": (PatternFill("solid", start_color="FFC7CE"), Font(color="9C0006", bold=True)),
    "std": (PatternFill("solid", start_color="F2F2F2"), Font(color="595959")),
}
FILL_SECTION = PatternFill("solid", start_color="D9E1F2")
MONEY = '"$"#,##0'
THIN_BOTTOM = Border(bottom=Side(style="thin", color="D0D0D0"))


def safe_sheet_name(title: str, taken: set[str]) -> str:
    name = (title or "Sheet")[:31]
    for ch in r"[]:*?/\\":
        name = name.replace(ch, " ")
    base, n = name, 2
    while name in taken:
        suffix = f" ({n})"
        name = base[: 31 - len(suffix)] + suffix
        n += 1
    taken.add(name)
    return name


def _fnum(v) -> float | None:
    try:
        return float(str(v).replace("$", "").replace(",", ""))
    except (TypeError, ValueError):
        return None


def _rule_matches(rule: dict, cell) -> bool:
    """A rule matches if ANY of its present tests pass. eq/contains are
    case-insensitive text tests (eq also matches numerically, so eq:5 hits
    '5.0'); gte/lte are numeric."""
    s = "" if cell is None else str(cell).strip().lower()
    if "eq" in rule:
        t = str(rule["eq"]).strip().lower()
        if s == t:
            return True
        a, b = _fnum(s), _fnum(t)
        if a is not None and b is not None and a == b:
            return True
    if "contains" in rule:
        needle = str(rule["contains"]).lower()
        if needle and needle in s:
            return True
    v = _fnum(s)
    if v is not None:
        g, l = _fnum(rule.get("gte")), _fnum(rule.get("lte"))
        if g is not None and v >= g:
            return True
        if l is not None and v <= l:
            return True
    return False


def tier_class(styling: dict | None, columns: list[dict], row: list) -> str | None:
    """First matching tier rule wins (same semantics as the frontend)."""
    if not styling:
        return None
    idx = {c["key"]: i for i, c in enumerate(columns)}
    for rule in styling.get("tier_rules", []):
        i = idx.get(rule.get("column"))
        if i is None or rule.get("class") not in FILLS:
            continue
        if _rule_matches(rule, row[i]):
            return rule["class"]
    return None


def write_table_sheet(wb: Workbook, name: str, spec: dict) -> None:
    ws = wb.create_sheet(name)
    columns = [c for c in spec["columns"] if not c.get("hidden")]
    keys = [c["key"] for c in spec["columns"]]
    vis_idx = [keys.index(c["key"]) for c in columns]

    for j, c in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=j, value=c.get("label") or c["key"])
        cell.fill = FILL_HDR
        cell.font = FONT_HDR
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"

    money_cols = {j for j, c in enumerate(columns, start=1)
                  if c.get("format") == "money"}
    # columns may point at a (hidden) sibling column holding the listing URL
    link_idx = {j: keys.index(c["link_col"])
                for j, c in enumerate(columns, start=1)
                if c.get("link_col") in keys}
    for i, row in enumerate(spec["rows"], start=2):
        klass = tier_class(spec.get("styling"), spec["columns"], row)
        for j, vi in enumerate(vis_idx, start=1):
            v = row[vi]
            url = row[link_idx[j]] if j in link_idx else None
            if not (isinstance(url, str) and url.startswith("http")):
                url = None
            if v in (None, "") and url:
                v = "verify"
            cell = ws.cell(row=i, column=j, value=v)
            cell.border = THIN_BOTTOM
            if j in money_cols and v not in (None, "") and v != "verify":
                cell.number_format = MONEY
            if url:
                cell.hyperlink = url
                cell.font = Font(color="0563C1", underline="single")
            if klass and klass in FILLS:
                cell.fill, cell.font = FILLS[klass]

    widths = {}
    wrap_cols = set()
    for j, c in enumerate(columns, start=1):
        sample = [str(c.get("label") or "")] + [
            str(r[vis_idx[j - 1]] or "") for r in spec["rows"][:50]]
        need = max(len(s) for s in sample) + 2
        widths[j] = min(46, max(10, need))
        if need > 46:
            wrap_cols.add(j)   # capped column -> wrap text like the UI does
    for j, w in widths.items():
        ws.column_dimensions[get_column_letter(j)].width = w
    for j in wrap_cols:
        for i in range(2, len(spec["rows"]) + 2):
            ws.cell(row=i, column=j).alignment = Alignment(
                wrap_text=True, vertical="top")
    if spec["rows"]:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{len(spec['rows']) + 1}"


def write_markdown_sheet(wb: Workbook, name: str, spec: dict) -> None:
    ws = wb.create_sheet(name)
    text = spec["rows"][0][0] if spec["rows"] else ""
    ws.column_dimensions["A"].width = 110
    for i, line in enumerate(str(text).split("\n"), start=1):
        cell = ws.cell(row=i, column=1, value=line.replace("#", "").strip())
        if line.startswith("## "):
            cell.font = Font(bold=True, size=13, color="1F4E79")
        elif line.startswith("**") or line.startswith("### "):
            cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="top")


def write_methodology(wb: Workbook, specs: list[dict]) -> None:
    """Generated purely from artifact provenance + warnings — no hand-written
    text to drift (the t11 convention: every workbook ends with sources)."""
    ws = wb.create_sheet("Methodology & Sources")
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 70
    ws.column_dimensions["C"].width = 40
    ws.column_dimensions["D"].width = 22
    for j, h in enumerate(("Source", "Detail", "URL", "Pulled"), start=1):
        cell = ws.cell(row=1, column=j, value=h)
        cell.fill = FILL_HDR
        cell.font = FONT_HDR
    seen: set[tuple] = set()
    r = 2
    for spec in specs:
        for p in spec.get("provenance") or []:
            key = (p.get("source"), p.get("url"))
            if key in seen:
                continue
            seen.add(key)
            ws.cell(row=r, column=1, value=p.get("source"))
            ws.cell(row=r, column=2, value=p.get("detail"))
            ws.cell(row=r, column=3, value=p.get("url"))
            ws.cell(row=r, column=4, value=(p.get("fetched_at") or "")[:10])
            r += 1
    r += 1
    cell = ws.cell(row=r, column=1, value="Caveats")
    cell.fill = FILL_SECTION
    cell.font = Font(bold=True)
    r += 1
    for note in _standing_caveats(specs):
        ws.cell(row=r, column=1, value="•")
        ws.cell(row=r, column=2, value=note)
        r += 1


def _standing_caveats(specs: list[dict]) -> list[str]:
    notes = {
        "CRDC science sections lag ~2 years (latest federal collection).",
    }
    # tool-emitted warnings are stored on the envelope, not the artifact; the
    # standing set covers the known structural caveats
    for spec in specs:
        for p in spec.get("provenance") or []:
            src = (p.get("source") or "").lower()
            if "checkbook" in src or "socrata" in src:
                notes.add("Open-checkbook coverage is partial: district-level "
                          "vendor spend exists only in DE, MD, NYC, Providence RI; "
                          "no rows elsewhere means NO DATA, not no purchases.")
            if "overpass" in src or "openstreetmap" in src:
                notes.add("OSM chemical/lab tagging is sparse — counts are a "
                          "floor, not a census.")
    return sorted(notes)


def _rule_legend(rule: dict) -> str:
    """'hot: Top target — 200k+ nearby' or 'hot: enroll_15mi gte 200000'."""
    klass = rule.get("class")
    if klass not in FILLS:
        return ""
    label = str(rule.get("label") or "").strip()
    if not label:
        test = next((f"{rule.get('column')} {k} {rule[k]}"
                     for k in ("gte", "lte", "eq", "contains") if k in rule),
                    str(rule.get("column") or ""))
        label = test
    return f"{klass}: {label}"


def write_readme(wb: Workbook, entries: list[tuple[str, str]]) -> None:
    ws = wb.create_sheet("READ ME", 0)
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 80
    c = ws.cell(row=1, column=1, value="SalesAgent export")
    c.font = Font(bold=True, size=14, color="1F4E79")
    ws.cell(row=2, column=1, value="Tab")
    ws.cell(row=2, column=2, value="What it is")
    for j in (1, 2):
        ws.cell(row=2, column=j).fill = FILL_HDR
        ws.cell(row=2, column=j).font = FONT_HDR
    r = 3
    for name, desc in entries:
        ws.cell(row=r, column=1, value=name)
        ws.cell(row=r, column=2, value=desc)
        r += 1
    r += 1
    ws.cell(row=r, column=1, value="Color legend").font = Font(bold=True)
    for klass, label in (("hot", "HOT (score >= 75 / act on it)"),
                         ("warm", "Warm (score >= 60)"),
                         ("now", "Act now / urgent window"),
                         ("std", "Standard")):
        r += 1
        cell = ws.cell(row=r, column=1, value=label)
        cell.fill, cell.font = FILLS[klass]


def build_workbook(specs: list[dict]) -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)
    taken: set[str] = {"READ ME", "Methodology & Sources"}
    entries = []
    for spec in specs:
        name = safe_sheet_name(spec.get("title") or spec["artifact_id"], taken)
        if spec["kind"] == "markdown":
            write_markdown_sheet(wb, name, spec)
        else:
            write_table_sheet(wb, name, spec)
        desc = (f"{spec['kind']} · {spec['row_count']} rows · "
                f"from {spec.get('created_by') or 'tool'}")
        rules = (spec.get("styling") or {}).get("tier_rules") or []
        keys = [r for r in (_rule_legend(rule) for rule in rules) if r]
        if keys:
            desc += " · colors: " + "; ".join(keys)
        entries.append((name, desc))
    write_readme(wb, entries)
    write_methodology(wb, specs)
    return wb
