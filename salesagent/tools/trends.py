"""TEMPORARY category-trend tools.

Data source = Category.xlsx dropped in the app root (a transposed CMT
Level-4 export: header rows Year/Quarter/Month, then per category x month two
measures — "Channel Cost (USD)" and "Organic Growth Channel Cost % (USD)").
Live file 2026-08-07: 145 categories, 2026-01..07. Category -> SKUs goes
through the Acumatica webstore feed by title keyword (crude on purpose).

Replace this module when a real per-SKU velocity source lands
(EC-DWH-SalesOrderLines in the GI catalog is the candidate).
"""
from __future__ import annotations

import difflib
import json
import re
from pathlib import Path

from ..db import utcnow
from ..integrations.acumatica import Acumatica, AcumaticaError
from .envelope import error_envelope, prov, table_envelope
from .registry import CostClass, tool_spec

_CACHE: dict = {"path": None, "mtime": None, "series": {}}
_STOP = {"and", "of", "the", "for", "with", "misc", "other"}
TRENDS_PTR = "trends_source.json"   # data/ pointer at the active uploaded file


def looks_like_category_export(header: list, body: list[list]) -> bool:
    """Cheap shape check for the CMT trends export (transposed sheet: 'Year'
    header row, 'CMT Level 4' label within the first rows of column 0). Pure
    — unit-tested; used by the upload route to auto-detect the file."""
    if not header or str(header[0]).strip().lower() != "year":
        return False
    return any(r and str(r[0] or "").strip().lower() == "cmt level 4"
               for r in body[:6])


def set_trends_source(settings, *, path, filename: str, upload_id: str,
                      conversation_id: str) -> None:
    (settings.data_dir / TRENDS_PTR).write_text(json.dumps({
        "path": str(path), "filename": filename, "upload_id": upload_id,
        "conversation_id": conversation_id, "set_at": utcnow()}),
        encoding="utf-8")


def resolve_trends_source(settings) -> tuple[Path, str, str | None]:
    """(path, human label, upload_id) — the most recently chat-uploaded
    category export wins; the app-root Category.xlsx is the fallback."""
    ptr = settings.data_dir / TRENDS_PTR
    if ptr.exists():
        try:
            d = json.loads(ptr.read_text(encoding="utf-8"))
            p = Path(d["path"])
            if p.exists():
                return (p, f"chat upload '{d.get('filename')}' "
                           f"({str(d.get('set_at', ''))[:10]})",
                        d.get("upload_id"))
        except Exception:
            pass
    return settings.category_xlsx, "app-root Category.xlsx", None


def load_category_series(path) -> dict[str, list[dict]]:
    """{category: [{month:'2026-01', cost: float|None, growth: frac|None},...]}
    (sorted by month). mtime-cached — the file is read per process, not per
    call."""
    mtime = path.stat().st_mtime
    if _CACHE["path"] == str(path) and _CACHE["mtime"] == mtime:
        return _CACHE["series"]
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Export"] if "Export" in wb.sheetnames else wb.worksheets[0]
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    wb.close()
    if len(rows) < 5:
        raise ValueError("Category.xlsx: unexpected shape (fewer than 5 rows)")
    years, months, measures = rows[0], rows[2], rows[3]
    cols = []
    for i in range(1, len(measures)):
        y = years[i] if i < len(years) else None
        m = months[i] if i < len(months) else None
        meas = measures[i]
        if not (y and m and meas):
            continue   # grand-total / blank columns
        kind = "growth" if "growth" in str(meas).lower() else "cost"
        cols.append((i, f"{str(y).strip()}-{str(m).strip()}", kind))
    series: dict[str, list[dict]] = {}
    for r in rows[4:]:
        cat = (str(r[0]).strip() if r and r[0] is not None else "")
        if not cat:
            continue
        by_m: dict[str, dict] = {}
        for i, ym, kind in cols:
            v = r[i] if i < len(r) else None
            if isinstance(v, (int, float)):
                by_m.setdefault(ym, {})[kind] = float(v)
        series[cat] = [{"month": ym, "cost": d.get("cost"),
                        "growth": d.get("growth")}
                       for ym, d in sorted(by_m.items())]
    _CACHE.update(path=str(path), mtime=mtime, series=series)
    return series


def trend_metrics(points: list[dict], n: int = 3) -> dict | None:
    """Pure — unit-tested. Growth values arrive as fractions; outputs are %."""
    pts = [p for p in points if p.get("cost") is not None]
    if not pts:
        return None
    last = pts[-1]
    lastn = pts[-n:]
    prevn = pts[-2 * n:-n]
    sum_last = sum(p["cost"] for p in lastn)
    sum_prev = sum(p["cost"] for p in prevn) if prevn else None
    momentum = (100.0 * (sum_last - sum_prev) / sum_prev
                if sum_prev else None)
    growths = [p["growth"] for p in lastn if p.get("growth") is not None]
    return {
        "latest_month": last["month"],
        "latest_cost": round(last["cost"], 2),
        "latest_growth_pct": (round(100 * last["growth"], 1)
                              if last.get("growth") is not None else None),
        "avg_growth_pct": (round(100 * sum(growths) / len(growths), 1)
                           if growths else None),
        "momentum_pct": round(momentum, 1) if momentum is not None else None,
    }


def category_keywords(category: str) -> tuple[str, list[str]]:
    """(server_keyword, score_words). Head noun of the category name, crudely
    singularized and stemmed to 7 chars so 'Pipetters' still hits
    'pipette'/'pipettor' titles. Pure — unit-tested."""
    words = [w for w in re.split(r"[^a-z]+", category.lower())
             if w and w not in _STOP]
    if not words:
        return "", []
    sing = [w[:-1] if w.endswith("s") and not w.endswith("ss") and len(w) > 3
            else w for w in words]
    primary = sing[-1]
    return primary[:7], sing


def _match_category(series: dict, name: str) -> str | None:
    low = name.strip().lower()
    for c in series:
        if c.lower() == low:
            return c
    part = [c for c in series if low in c.lower() or c.lower() in low]
    if len(part) == 1:
        return part[0]
    close = difflib.get_close_matches(name, list(series), n=1, cutoff=0.75)
    return close[0] if close else None


@tool_spec(
    name="category_trends",
    description=(
        "TEMPORARY: what's doing well right now, by product category — "
        "monthly channel cost (USD) + organic growth % per CMT category. "
        "Source = the most recent category export the user DROPPED INTO CHAT "
        "(auto-detected on upload), else the app-root Category.xlsx "
        "(category level only, NO per-SKU velocity). Default = "
        "ranked table (latest cost, latest & avg growth %, momentum = last N "
        "months vs the N before). Pass category='...' for that category's "
        "per-month detail (chartable). Follow up with category_skus for "
        "in-stock SKUs to recommend. Free."),
    input_schema={"properties": {
        "months_back": {"type": "integer", "default": 3,
                        "description": "window for avg growth + momentum"},
        "min_monthly_cost": {"type": "number", "default": 1000,
                             "description": "floor for riser/decliner callouts"},
        "category": {"type": "string",
                     "description": "one category's per-month detail"},
    }},
    cost_class=CostClass.FREE,
)
def category_trends(ctx, months_back: int = 3, min_monthly_cost: float = 1000,
                    category: str | None = None) -> dict:
    path, src_label, _uid = resolve_trends_source(ctx.settings)
    if not path.exists():
        return error_envelope(
            "no trend source: drop a Category export into the chat (it's "
            "auto-detected) or place Category.xlsx in the app root",
            error_type="NotConfigured")
    try:
        series = load_category_series(path)
    except Exception as e:
        return error_envelope(
            f"could not parse the trends file ({src_label}): {e}")
    n = max(1, min(int(months_back or 3), 6))
    src = prov(f"CMT Level-4 channel-cost export — {src_label} (temp source)",
               f"{len(series)} categories; file: {path.name}")

    if category:
        cat = _match_category(series, category)
        if not cat:
            close = difflib.get_close_matches(category, list(series), n=5,
                                              cutoff=0.4)
            return error_envelope(
                f"no category '{category}'; closest: {close}")
        rows = [[p["month"], p["cost"],
                 round(100 * p["growth"], 1) if p.get("growth") is not None
                 else None]
                for p in series[cat]]
        return table_envelope(
            ctx.rw(), ctx.emit, conversation_id=ctx.conversation_id,
            tool="category_trends", title=f"Trend — {cat}",
            columns=[{"key": "month", "label": "Month", "type": "string"},
                     {"key": "cost", "label": "Channel cost $", "type": "number",
                      "format": "money"},
                     {"key": "growth_pct", "label": "Organic growth %",
                      "type": "number"}],
            rows=rows, provenance=[src],
            summary=f"{cat}: {len(rows)} months of channel cost + growth.")

    rows = []
    for cat, pts in series.items():
        m = trend_metrics(pts, n)
        if m:
            rows.append([cat, m["latest_cost"], m["latest_growth_pct"],
                         m["avg_growth_pct"], m["momentum_pct"],
                         m["latest_month"]])
    total_latest = sum(r[1] for r in rows if r[1]) or 1.0
    for r in rows:
        r.append(round(100 * (r[1] or 0) / total_latest, 1))
    rows.sort(key=lambda r: (r[3] is None, -(r[3] or 0)))
    big = [r for r in rows if (r[1] or 0) >= float(min_monthly_cost or 0)]
    risers = [f"{r[0]} ({r[3]:+.0f}% avg)" for r in big[:3]
              if r[3] is not None]
    decliners = [f"{r[0]} ({r[3]:+.0f}% avg)" for r in reversed(big[-3:])
                 if r[3] is not None]
    cols = [{"key": "category", "label": "Category", "type": "string"},
            {"key": "latest_cost", "label": "Latest month $", "type": "number",
             "format": "money"},
            {"key": "latest_growth_pct", "label": "Latest growth %",
             "type": "number"},
            {"key": "avg_growth_pct", "label": f"Avg growth % ({n}mo)",
             "type": "number"},
            {"key": "momentum_pct", "label": f"Momentum % ({n}mo vs prior)",
             "type": "number"},
            {"key": "latest_month", "label": "As of", "type": "string"},
            {"key": "share_pct", "label": "Share %", "type": "number"}]
    return table_envelope(
        ctx.rw(), ctx.emit, conversation_id=ctx.conversation_id,
        tool="category_trends",
        title=f"Category trends (avg growth, last {n} mo)",
        columns=cols, rows=rows, provenance=[src],
        summary=(f"{len(rows)} categories through "
                 f"{rows[0][5] if rows else '?'} (source: {src_label}). "
                 f"Top risers (≥${min_monthly_cost:g}/mo): {', '.join(risers)}. "
                 f"Biggest decliners: {', '.join(decliners)}."),
        warnings=["category-level channel cost only — no per-SKU velocity in "
                  "this source (temporary wiring)",
                  "growth % in the hundreds+ usually means a near-zero base "
                  "month — sanity-check against latest-month $ and momentum "
                  "before calling it a trend"])


@tool_spec(
    name="category_skus",
    description=(
        "TEMPORARY companion to category_trends: in-stock Eisco SKUs for a "
        "category, matched by TITLE KEYWORD against the Acumatica webstore "
        "feed (crude — pass keyword='...' to override the derived one; the "
        "summary echoes what was searched). Returns sku/title/brand/price/"
        "webstore stock, in-stock first. Verify exact availability with "
        "sku_stock before promising. Free."),
    input_schema={"properties": {
        "category": {"type": "string"},
        "keyword": {"type": "string",
                    "description": "override the derived title keyword"},
        "in_stock_only": {"type": "boolean", "default": True},
        "limit": {"type": "integer", "default": 25, "maximum": 100},
    }, "required": ["category"]},
    cost_class=CostClass.FREE,
)
def category_skus(ctx, category: str, keyword: str | None = None,
                  in_stock_only: bool = True, limit: int = 25) -> dict:
    acu = Acumatica(ctx.settings)
    if not acu.configured:
        return error_envelope("direct Acumatica not configured",
                              error_type="NotConfigured")
    kw, score_words = ((keyword.strip().lower(), [keyword.strip().lower()])
                       if keyword else category_keywords(category))
    if not kw:
        return error_envelope(f"could not derive a keyword from '{category}' — "
                              "pass keyword='...'")
    try:
        prods = acu.products_by_title(kw)
    except AcumaticaError as e:
        return error_envelope(str(e), error_type="AcumaticaError")
    for p in prods:
        tl = p["title"].lower()
        p["score"] = sum(1 for w in score_words if w in tl)
    prods.sort(key=lambda p: (-(p["stock"] > 0), -p["score"], -p["stock"]))
    if in_stock_only:
        kept = [p for p in prods if p["stock"] > 0]
    else:
        kept = prods
    kept = kept[: max(1, min(int(limit or 25), 100))]
    rows = [[p["sku"], p["title"], p["brand"], p["price"], p["stock"]]
            for p in kept]
    cols = [{"key": "sku", "label": "SKU", "type": "string"},
            {"key": "title", "label": "Product", "type": "string"},
            {"key": "brand", "label": "Brand", "type": "string"},
            {"key": "price", "label": "Webstore $", "type": "number",
             "format": "money"},
            {"key": "stock", "label": "Webstore stock", "type": "number"}]
    return table_envelope(
        ctx.rw(), ctx.emit, conversation_id=ctx.conversation_id,
        tool="category_skus",
        title=f"In-stock SKUs — {category}" if in_stock_only
        else f"SKUs — {category}",
        columns=cols, rows=rows,
        provenance=[prov("Acumatica webstore feed (EC-IPProducts, read-only)",
                         f"title contains '{kw}'; stock = webstore quantity")],
        summary=(f"{len(rows)} SKUs for '{category}' via title keyword '{kw}' "
                 f"({sum(1 for p in prods if p['stock'] > 0)} of "
                 f"{len(prods)} matches have webstore stock). Keyword "
                 "matching is crude — override with keyword='...' if these "
                 "look wrong."),
        warnings=["category->SKU mapping is TITLE-KEYWORD based (temporary); "
                  "verify availability with sku_stock before promising"])
