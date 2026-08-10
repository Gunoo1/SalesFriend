"""File uploads: CSV/XLSX -> table artifact. The agent is told about it on
the next turn via drain_upload_notices() (routes_chat), and any list-taking
tool then reads values server-side by {artifact_id, column}."""
from __future__ import annotations

import csv
import io
import sqlite3
import uuid

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile

from ..artifacts import store
from ..auth import current_user, db_conn
from ..db import utcnow
from ..settings import load_settings

router = APIRouter(prefix="/api")

MAX_UPLOAD_BYTES = 20 * 1024 * 1024


def _parse_csv(data: bytes) -> tuple[list[str], list[list]]:
    for enc in ("utf-8-sig", "cp1252"):
        try:
            text = data.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    else:
        raise HTTPException(400, "could not decode file as UTF-8 or cp1252")
    sample = text[:4000]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel
    reader = csv.reader(io.StringIO(text), dialect)
    rows = [r for r in reader if any(str(c).strip() for c in r)]
    if not rows:
        raise HTTPException(400, "file is empty")
    header, body = rows[0], rows[1:]
    return [str(h).strip() or f"col{i+1}" for i, h in enumerate(header)], body


def _parse_xlsx(data: bytes) -> tuple[list[str], list[list]]:
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.worksheets[0]
    rows = [[c for c in row] for row in ws.iter_rows(values_only=True)]
    rows = [r for r in rows if any(c not in (None, "") for c in r)]
    wb.close()
    if not rows:
        raise HTTPException(400, "sheet is empty")
    header = [str(h).strip() if h not in (None, "") else f"col{i+1}"
              for i, h in enumerate(rows[0])]
    return header, [list(r) for r in rows[1:]]


@router.post("/conversations/{cid}/upload")
async def upload_file(cid: str, file: UploadFile = File(...),
                      user: dict = Depends(current_user),
                      conn: sqlite3.Connection = Depends(db_conn)):
    convo = conn.execute("SELECT * FROM conversations WHERE id=?",
                         (cid,)).fetchone()
    if not convo or convo["user_id"] != user["id"]:
        raise HTTPException(404, "conversation not found")
    data = await file.read()
    if len(data) > MAX_UPLOAD_BYTES:
        raise HTTPException(413, "file too large (20 MB max)")
    name = file.filename or "upload"
    lower = name.lower()
    if lower.endswith((".xlsx", ".xls")):
        header, body = _parse_xlsx(data)
    elif lower.endswith(".csv") or lower.endswith(".txt"):
        header, body = _parse_csv(data)
    else:
        raise HTTPException(400, "upload a .csv or .xlsx file")

    settings = load_settings()
    uid = uuid.uuid4().hex[:10]
    raw_path = settings.uploads_dir / f"{uid}_{name}"
    raw_path.write_bytes(data)

    # a dropped-in category-trends export becomes THE live trends source
    trends_source = False
    if lower.endswith((".xlsx", ".xls")):
        from ..tools.trends import looks_like_category_export, set_trends_source
        if looks_like_category_export(header, body):
            set_trends_source(settings, path=raw_path, filename=name,
                              upload_id=uid, conversation_id=cid)
            trends_source = True

    width = len(header)
    norm_rows = [(r + [None] * width)[:width] for r in body]
    columns = [{"key": h, "label": h, "type": "string"} for h in header]
    art = store.create(
        conn, conversation_id=cid, tool="upload", kind="table",
        title=f"Upload — {name}", columns=columns, rows=norm_rows,
        provenance=[{"source": "user upload", "detail": name,
                     "url": None, "fetched_at": utcnow(),
                     "uploaded_by": user["username"]}])
    with conn:
        conn.execute(
            "INSERT INTO uploads (id, user_id, conversation_id, filename,"
            " artifact_id, path, created_at, notified) VALUES (?,?,?,?,?,?,?,0)",
            (uid, user["id"], cid, name, art["artifact_id"], str(raw_path),
             utcnow()))
        conn.execute("UPDATE conversations SET updated_at=? WHERE id=?",
                     (utcnow(), cid))
    return {"upload_id": uid, "artifact_id": art["artifact_id"],
            "version": art["version"], "kind": "table",
            "title": f"Upload — {name}", "filename": name,
            "rows": len(norm_rows), "columns": header,
            "trends_source": trends_source}


def drain_job_notices(conn: sqlite3.Connection, cid: str) -> str:
    """Finished-but-unnotified jobs -> [system] block for the agent's next
    turn (mirrors upload notices)."""
    rows = conn.execute(
        "SELECT id, tool_name, status, result_artifact_id, error FROM jobs "
        "WHERE conversation_id=? AND notified=0 AND status IN "
        "('done','error') ORDER BY finished_at", (cid,)).fetchall()
    if not rows:
        return ""
    notes = []
    for r in rows:
        if r["status"] == "done":
            notes.append(f"[system] background job {r['id']} ({r['tool_name']}) "
                         f"finished -> artifact {r['result_artifact_id']}. "
                         "Reference or transform it as needed.")
        else:
            notes.append(f"[system] background job {r['id']} ({r['tool_name']}) "
                         f"FAILED: {r['error']}")
    with conn:
        conn.execute(
            f"UPDATE jobs SET notified=1 WHERE id IN "
            f"({','.join('?' * len(rows))})", [r["id"] for r in rows])
    return "\n".join(notes)


def drain_upload_notices(conn: sqlite3.Connection, cid: str) -> str:
    """Unnotified uploads -> one [system] block prepended to the agent's view
    of the next user message (the persisted message stays clean)."""
    rows = conn.execute(
        "SELECT u.id, u.filename, u.artifact_id FROM uploads u "
        "WHERE u.conversation_id=? AND u.notified=0 ORDER BY u.created_at",
        (cid,)).fetchall()
    if not rows:
        return ""
    trends_uid = None
    try:
        from ..tools.trends import resolve_trends_source
        _, _, trends_uid = resolve_trends_source(load_settings())
    except Exception:
        pass
    notes = []
    for r in rows:
        if r["id"] == trends_uid:
            notes.append(
                f"[system] The user uploaded '{r['filename']}' — recognized "
                "as a CATEGORY-TRENDS export; category_trends now reads THIS "
                "file directly (skip artifact_peek for trend questions).")
            continue
        spec_row = conn.execute(
            "SELECT columns_json, row_count FROM artifact_versions "
            "WHERE artifact_id=? AND version=1", (r["artifact_id"],)).fetchone()
        cols = ""
        n = "?"
        if spec_row:
            import json
            cols = ", ".join(c["key"] for c in
                             json.loads(spec_row["columns_json"])[:12])
            n = spec_row["row_count"]
        notes.append(f"[system] The user uploaded '{r['filename']}' -> artifact "
                     f"{r['artifact_id']} ({n} rows; columns: {cols}). Read "
                     "values with artifact_peek; list-taking tools accept "
                     "{artifact_id, column}.")
    with conn:
        conn.execute(
            f"UPDATE uploads SET notified=1 WHERE id IN "
            f"({','.join('?' * len(rows))})", [r["id"] for r in rows])
    return "\n".join(notes)
